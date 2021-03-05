import pandas as pd
from openpyxl import load_workbook

file_location = r"C:\Users\S3017862\OneDrive\Documents\BBB\E-brewR2.4_ADDIS PALE_BREW41-42_211070_01.xlsm"
brewlog_df = pd.read_excel(file_location, sheet_name="BREWLOG",engine='openpyxl')

existing_database_file = 'BBBlog.xlsx'


#%%

cell_locations = {
    'brew_name': (1,1),
    'brew_type': (2,1),
    'recipe_ver': (2,4),
    'date': (4,4),
    'length': (3,1),
    'brew_no': (4,1),
    'temp_amb_c': (5,1),
    'vol': (7,1),
    'digi_temp_c': (7,3),
    'mash_start_time': (9,2),
    'mash_all_in_time': (9,6),
    'mash_2min_temp': (12,4),
    'mash_2min_ph': (12,5),
    'mash_2min_sg': (12,6),
    'mash_2min_time': (12,7),
    'mash_15min_temp': (15, 4),
    'mash_15min_ph': (15, 5),
    'mash_15min_sg': (15, 6),
    'mash_15min_time': (15, 7),
    'mash_60min_temp': (18, 4),
    'mash_60min_ph': (18, 5),
    'mash_60min_sg': (18, 6),
    'mash_60min_time': (18, 7),
    'fist_runnings_temp': (31,4),
    'fist_runnings_ph': (31, 5),
    'fist_runnings_sg': (31, 6),
    'fist_runnings_time': (31, 7),
    'fist_runnings_cloudy': (32, 6),
    'fist_runnings_clear': (33, 6),
    'fist_runnings_brilliant': (34, 6),
    'final_runnings_temp': (36, 4),
    'final_runnings_ph': (36, 5),
    'final_runnings_sg': (36, 6),
    'final_runnings_time': (36, 7),
    'final_runnings_cloudy': (37, 6),
    'final_runnings_clear': (38, 6),
    'final_runnings_brilliant': (39, 6),
    'runnings_time': (41,7),
    'pre_boil_vol': (43,1),
    'post_boil_vol': (45,1),
    'post_boil_ph': (45,5),
    'post_boil_sg': (45,6),
    'vol_into_fermenter': (51,1),
    'liquir_back_vol': (53,1),
    'fv_number': (51,5),
    'fermenter_total_vol': (55,1),
    'fermenter_ph': (55, 5),
    'fermenter_sg': (55, 6),
    'fermenter_o2ppm': (55, 7)
}

#%% create dictionary from cell locations to create new line in the form of a workbook

cell_values = { key: [brewlog_df.iloc[val]] for key,val in cell_locations.items()}
new_line = pd.DataFrame(cell_values)

#%%

book = load_workbook(existing_database_file)
writer = pd.ExcelWriter(existing_database_file, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
startrow = book.active.max_row
new_line.to_excel(writer, startrow = startrow, index=False, header=False)
writer.close()